import io
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill

EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".xlsb"}


def apply_app_styles() -> None:
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Barlow:wght@400;500;600;700&display=swap');

        .stApp {
            background:
                radial-gradient(circle at 0% 0%, #dbeafe 0, rgba(219, 234, 254, 0) 40%),
                radial-gradient(circle at 100% 100%, #d1fae5 0, rgba(209, 250, 229, 0) 38%),
                linear-gradient(180deg, #f8fafc 0%, #eef2f7 100%);
            font-family: 'Barlow', sans-serif;
        }
        .block-container {
            max-width: 1100px;
            padding-top: 1.25rem;
            padding-bottom: 2.25rem;
        }
        .hero-card {
            background: linear-gradient(135deg, #0f4c81 0%, #1d4ed8 100%);
            border-radius: 18px;
            padding: 1.1rem 1.2rem;
            color: #ffffff;
            margin-bottom: 1rem;
            box-shadow: 0 14px 30px rgba(15, 76, 129, 0.25);
        }
        .hero-title {
            font-size: 1.45rem;
            font-weight: 700;
            letter-spacing: 0.2px;
        }
        .hero-subtitle {
            margin-top: 0.3rem;
            font-size: 0.97rem;
            color: #e2e8f0;
        }
        .section-title {
            margin: 0.35rem 0 0.65rem 0;
            color: #0f172a;
            font-weight: 700;
            font-size: 1.03rem;
        }
        .status-wrap {
            display: flex;
            gap: 0.45rem;
            flex-wrap: wrap;
            margin-top: 0.2rem;
            margin-bottom: 0.7rem;
        }
        .status-pill {
            border-radius: 999px;
            padding: 0.25rem 0.68rem;
            font-size: 0.82rem;
            font-weight: 600;
            border: 1px solid transparent;
        }
        .status-ok {
            background: #dcfce7;
            border-color: #86efac;
            color: #166534;
        }
        .status-warn {
            background: #fff7ed;
            border-color: #fed7aa;
            color: #9a3412;
        }
        [data-testid="stTextInputRootElement"] > div,
        [data-testid="stFileUploader"] section,
        [data-testid="stDataFrame"] {
            background: rgba(255, 255, 255, 0.86);
            border: 1px solid #dbe2ea;
            border-radius: 14px;
        }
        [data-testid="stTextInputRootElement"] > div,
        [data-testid="stFileUploader"] section {
            box-shadow: 0 6px 18px rgba(15, 23, 42, 0.05);
        }
        [data-testid="stFileUploader"] small {
            color: #475569;
        }
        [data-testid="stButton"] button,
        [data-testid="stDownloadButton"] button {
            border-radius: 12px;
            border: 0;
            background: linear-gradient(135deg, #0f4c81 0%, #1d4ed8 100%);
            color: #ffffff;
            font-weight: 600;
            min-height: 2.75rem;
            box-shadow: 0 10px 22px rgba(29, 78, 216, 0.25);
        }
        [data-testid="stButton"] button:hover,
        [data-testid="stDownloadButton"] button:hover {
            filter: brightness(1.05);
        }
        [data-testid="stDataFrame"] {
            border-radius: 14px;
            overflow: hidden;
        }
        [data-testid="stAlert"] {
            border-radius: 12px;
        }
        h3 {
            color: #0f172a;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def normalize_key(value: object) -> str:
    return "" if pd.isna(value) else str(value).strip().lower()


def normalize_position_key(value: object) -> str:
    return "" if pd.isna(value) else str(value).strip().lower()


def first_variant_size_by_color(dataframe: pd.DataFrame) -> dict[str, str]:
    sizes: dict[str, str] = {}
    for color, group_df in dataframe.groupby("Color", sort=False):
        variant_rows = group_df[
            group_df["Operation"].fillna("").astype(str).str.strip().str.lower().str.startswith("variant")
        ]
        if variant_rows.empty:
            sizes[str(color)] = ""
        else:
            first_size = variant_rows.iloc[0]["Size"]
            sizes[str(color)] = "" if pd.isna(first_size) else str(first_size)
    return sizes


def count_size_values(size_value: str) -> int:
    if not size_value:
        return 0
    parts = [part.strip() for part in str(size_value).split(",") if part.strip()]
    return len(parts)


def count_items_by_color(dataframe: pd.DataFrame) -> dict[str, int]:
    counts: dict[str, int] = {}
    for color, group_df in dataframe.groupby("Color", sort=False):
        operations = group_df["Operation"].fillna("").astype(str).str.strip().tolist()
        item_count = 0
        for operation in operations:
            normalized = operation.lower()
            if normalized.startswith("production"):
                continue
            if normalized.startswith(("r", "s", "p")):
                item_count += 1
        counts[str(color)] = item_count
    return counts


def parse_operation_fields(operation_value: object) -> pd.Series:
    operation_text = "" if pd.isna(operation_value) else str(operation_value).strip()
    if not operation_text:
        return pd.Series({
            "PLM No": "",
            "Col": "",
            "SAP No.": "",
            "Width": "",
            "Position": "",
        })

    tokens = operation_text.split()
    if not tokens or tokens[0][0] not in {"R", "S", "P"}:
        return pd.Series({
            "PLM No": "",
            "Col": "",
            "SAP No.": "",
            "Width": "",
            "Position": "",
        })

    prefix = tokens[0][0]
    plm_no = tokens[0]

    def is_sap_token(token: str) -> bool:
        return token.isdigit() and len(token) >= 7

    sap_index = None
    for index, token in enumerate(tokens[1:], start=1):
        if is_sap_token(token):
            sap_index = index
            break

    col_value = ""
    if len(tokens) > 1 and sap_index not in (None, 1):
        col_value = tokens[1]

    sap_value = tokens[sap_index] if sap_index is not None else ""

    width_value = ""
    position_value = ""
    if sap_index is not None and prefix in {"R", "S"}:
        remainder_tokens = tokens[sap_index + 1 :]
        width_index = None
        for index in range(len(remainder_tokens) - 1, -1, -1):
            token = remainder_tokens[index]
            if token.isdigit():
                width_index = index
                break

        if width_index is not None:
            width_value = remainder_tokens[width_index]
            position_value = " ".join(remainder_tokens[width_index + 1 :]).strip()

    return pd.Series({
        "PLM No": plm_no,
        "Col": col_value,
        "SAP No.": sap_value,
        "Width": width_value,
        "Position": position_value,
    })


def format_output_sheet(
    worksheet,
    highlight_numeric_operation: bool = False,
    freeze_top_row: bool = False,
    highlight_plm_check_header: bool = False,
    highlight_operation_check_yes: bool = False,
    highlight_row_column_name: str = "",
    highlight_row_match_value: str = "",
) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    plm_check_header_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    highlight_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    header_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    if highlight_plm_check_header:
        for cell in worksheet[1]:
            if str(cell.value).strip().lower() == "plm check":
                cell.fill = plm_check_header_fill
                break

    if highlight_numeric_operation:
        operation_column = None
        for cell in worksheet[1]:
            if cell.value == "Operation":
                operation_column = cell.column
                break

        if operation_column is not None:
            for row in range(2, worksheet.max_row + 1):
                operation_cell = worksheet.cell(row=row, column=operation_column)
                operation_text = "" if operation_cell.value is None else str(operation_cell.value).strip()
                if operation_text and operation_text[0].isdigit():
                    for column in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=column).fill = highlight_fill

    if highlight_operation_check_yes:
        operation_check_column = None
        for cell in worksheet[1]:
            if str(cell.value).strip().lower() == "operation check":
                operation_check_column = cell.column
                break

        if operation_check_column is not None:
            for row in range(2, worksheet.max_row + 1):
                operation_check_cell = worksheet.cell(row=row, column=operation_check_column)
                operation_check_text = (
                    "" if operation_check_cell.value is None else str(operation_check_cell.value).strip().lower()
                )
                if operation_check_text == "yes":
                    for column in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=column).fill = highlight_fill

    if highlight_row_column_name and highlight_row_match_value:
        target_column = None
        for cell in worksheet[1]:
            if str(cell.value).strip().lower() == highlight_row_column_name.strip().lower():
                target_column = cell.column
                break

        if target_column is not None:
            match_value_normalized = highlight_row_match_value.strip().lower()
            for row in range(2, worksheet.max_row + 1):
                target_cell = worksheet.cell(row=row, column=target_column)
                target_text = "" if target_cell.value is None else str(target_cell.value).strip().lower()
                if target_text == match_value_normalized:
                    for column in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=column).fill = highlight_fill

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    if freeze_top_row:
        worksheet.freeze_panes = "A2"


def build_master_from_uploads(uploaded_files: list[Any], style_name: str) -> tuple[pd.DataFrame, dict[str, dict[str, Any]], list[str]]:
    logs: list[str] = []
    merged_frames: list[pd.DataFrame] = []

    for file in uploaded_files:
        file_name = Path(file.name).name
        suffix = Path(file_name).suffix.lower()
        if suffix not in EXCEL_EXTENSIONS:
            logs.append(f"Skipped non-Excel file: {file_name}")
            continue

        try:
            color_name = Path(file_name).stem
            if "-" in color_name:
                color_name = color_name.split("-", 1)[1].strip()

            raw_df = pd.read_excel(file, usecols="A:D", header=None)
            raw_df = raw_df.dropna(how="all")

            if raw_df.empty:
                logs.append(f"Skipped empty file: {file_name}")
                continue

            first_row = [str(v).strip().lower() for v in raw_df.iloc[0].tolist()]
            if first_row[:4] in (
                ["option", "size", "description", "needle"],
                ["operation", "size", "description", "needle"],
            ):
                raw_df = raw_df.iloc[1:]

            if raw_df.empty:
                logs.append(f"Skipped header-only file: {file_name}")
                continue

            transformed_df = pd.DataFrame(
                {
                    "Style": style_name,
                    "Color": color_name,
                    "Operation": raw_df.iloc[:, 0],
                    "Size": raw_df.iloc[:, 1],
                    "Description": raw_df.iloc[:, 2],
                    "Needle": raw_df.iloc[:, 3],
                }
            )
            transformed_df = transformed_df.dropna(subset=["Operation", "Size", "Description", "Needle"], how="all")

            if transformed_df.empty:
                logs.append(f"Skipped blank rows file: {file_name}")
                continue

            merged_frames.append(transformed_df)
            logs.append(f"Processed: {file_name} ({len(transformed_df)} rows)")
        except Exception as file_error:
            logs.append(f"File error ({file_name}): {file_error}")

    if not merged_frames:
        raise ValueError("No valid rows found in uploaded files.")

    merged_df = pd.concat(merged_frames, ignore_index=True)
    item_counts = count_items_by_color(merged_df)
    variant_sizes = first_variant_size_by_color(merged_df)

    summary: dict[str, dict[str, Any]] = {}
    for color in merged_df["Color"].astype(str).drop_duplicates().tolist():
        size_value = variant_sizes.get(color, "")
        summary[color] = {
            "Item Count": item_counts.get(color, 0),
            "Size": size_value,
            "#No Size": count_size_values(size_value),
        }

    return merged_df, summary, logs


def generate_output_workbook(merged_data: pd.DataFrame, style_name: str, summary: dict[str, dict[str, Any]], plm_file: Any = None) -> tuple[bytes, dict[str, pd.DataFrame], list[str]]:
    logs: list[str] = []

    plm_df = None
    plm_lookup_df = None
    plm_no_index = None
    plm_position_index = None
    plm_consumption_index = None
    plm_base_unit_index = None

    if plm_file is not None:
        try:
            plm_df = pd.read_excel(plm_file, sheet_name=0, header=None)
            if not plm_df.empty and plm_df.shape[1] >= 2:
                plm_lookup_df = plm_df.copy()
                plm_lookup_df.iat[0, 1] = "PLM No"
                headers = [normalize_key(v) for v in plm_lookup_df.iloc[0].tolist()]

                for index, header in enumerate(headers):
                    if header in {"plm no", "plm_no", "plmno", "plm number"}:
                        plm_no_index = index
                    if header in {"position", "pos"}:
                        plm_position_index = index
                    if header in {"consumption", "cons", "consump"}:
                        plm_consumption_index = index
                    if header in {"base unit", "base_unit", "uom", "unit"}:
                        plm_base_unit_index = index

                if plm_no_index is not None and plm_position_index is not None:
                    logs.append("PLM columns detected: PLM No and Position")
                else:
                    logs.append("PLM lookup warning: Could not find PLM No/Position columns in PLM sheet.")

                if plm_consumption_index is not None and plm_base_unit_index is not None:
                    logs.append("PLM columns detected: Consumption and Base Unit")
                else:
                    logs.append("PLM lookup warning: Could not find Consumption/Base Unit columns in PLM sheet.")
            else:
                logs.append("PLM lookup warning: PLM file is empty or missing column B.")
        except Exception as plm_read_error:
            logs.append(f"PLM read error: {plm_read_error}")

    item_calc_df = pd.DataFrame(
        [
            {
                "Color": color,
                "Item Count": info["Item Count"],
                "Size": info["Size"],
                "#No Size": info["#No Size"],
            }
            for color, info in summary.items()
        ]
    )

    selected_color = ""
    if summary:
        max_no_size = max(info.get("#No Size", 0) for info in summary.values())
        top_no_size_colors = [color for color, info in summary.items() if info.get("#No Size", 0) == max_no_size]

        max_item_count = max(summary[color].get("Item Count", 0) for color in top_no_size_colors)
        top_item_count_colors = [
            color for color in top_no_size_colors if summary[color].get("Item Count", 0) == max_item_count
        ]

        selected_color = top_item_count_colors[0]
        selected_info = summary[selected_color]

        if len(top_no_size_colors) > 1:
            logs.append(
                f"Multiple colors have max #No Size ({max_no_size}). Compared Item Count; selected: {selected_color}"
            )
        if len(top_item_count_colors) > 1:
            logs.append("Multiple colors tied on #No Size and Item Count. " f"Selected: {selected_color}")

        logs.append(
            f"Combine source color: {selected_color} "
            f"(#No Size = {selected_info.get('#No Size', 0)}, Item Count = {selected_info.get('Item Count', 0)})"
        )

    if selected_color:
        master_color_series = merged_data["Color"].fillna("").astype(str)
        combine_source_df = merged_data[master_color_series == selected_color].copy()
    else:
        combine_source_df = merged_data.copy()

    if combine_source_df.empty:
        logs.append("No rows found for selected color in Master. Using all Master rows for Combine.")
        combine_source_df = merged_data.copy()

    combine_keys = ["Operation", "Size", "Description", "Needle"]
    combine_df = combine_source_df[["Color", *combine_keys]].reset_index(drop=True).copy()
    combine_operation_is_numeric = pd.to_numeric(combine_df["Operation"], errors="coerce").notna()
    combine_operation_name_series = combine_df["Operation"].where(combine_operation_is_numeric).ffill()
    combine_operation_name_series = combine_operation_name_series.apply(lambda value: "" if pd.isna(value) else str(value))
    combine_df.insert(1, "Operation Name", combine_operation_name_series)

    parsed_operation_df = combine_df["Operation"].apply(parse_operation_fields)
    combine_df["PLM No"] = parsed_operation_df["PLM No"]
    combine_df["Col"] = parsed_operation_df["Col"]
    combine_df["SAP No."] = parsed_operation_df["SAP No."]
    combine_df["Width"] = parsed_operation_df["Width"]
    combine_df["Consumption"] = ""
    combine_df["Base Unit"] = ""
    combine_df["Legacy"] = ""
    combine_df["Position"] = parsed_operation_df["Position"]

    combine_columns = list(combine_df.columns)
    if "Width" in combine_columns:
        width_index = combine_columns.index("Width")
        for column_name in ["Consumption", "Base Unit"]:
            if column_name in combine_columns:
                combine_columns.remove(column_name)
        combine_columns[width_index + 1 : width_index + 1] = ["Consumption", "Base Unit"]
        combine_df = combine_df[combine_columns]

    if (
        plm_lookup_df is not None
        and plm_no_index is not None
        and plm_position_index is not None
        and plm_consumption_index is not None
        and plm_base_unit_index is not None
    ):
        plm_value_lookup: dict[tuple[str, str], tuple[str, str]] = {}
        plm_value_by_plm_no: dict[str, tuple[str, str]] = {}
        for _, row in plm_lookup_df.iloc[1:].iterrows():
            plm_key = normalize_key(row.iloc[plm_no_index])
            position_key = normalize_position_key(row.iloc[plm_position_index])
            if not plm_key:
                continue

            consumption_value = "" if pd.isna(row.iloc[plm_consumption_index]) else str(row.iloc[plm_consumption_index]).strip()
            base_unit_value = "" if pd.isna(row.iloc[plm_base_unit_index]) else str(row.iloc[plm_base_unit_index]).strip()
            plm_value_lookup[(plm_key, position_key)] = (consumption_value, base_unit_value)

            if plm_key not in plm_value_by_plm_no:
                plm_value_by_plm_no[plm_key] = (consumption_value, base_unit_value)

        if plm_value_lookup or plm_value_by_plm_no:
            for row_index in combine_df.index:
                plm_key = normalize_key(combine_df.at[row_index, "PLM No"])
                position_key = normalize_position_key(combine_df.at[row_index, "Position"])
                if not plm_key:
                    continue

                if position_key:
                    match_values = plm_value_lookup.get((plm_key, position_key))
                else:
                    match_values = plm_value_by_plm_no.get(plm_key)

                if match_values is not None:
                    combine_df.at[row_index, "Consumption"] = match_values[0]
                    combine_df.at[row_index, "Base Unit"] = match_values[1]

    master_df = merged_data.copy()
    operation_is_numeric = pd.to_numeric(master_df["Operation"], errors="coerce").notna()
    operation_name_series = master_df["Operation"].where(operation_is_numeric).ffill()
    operation_name_series = operation_name_series.apply(lambda value: "" if pd.isna(value) else str(value))

    master_df.insert(0, "Operation Check", operation_is_numeric.map({True: "Yes", False: "No"}))
    master_df.insert(1, "Operation Name", operation_name_series)
    master_parsed_df = master_df["Operation"].apply(parse_operation_fields)
    master_df.insert(5, "PLM No", master_parsed_df["PLM No"])
    master_df.insert(6, "Col", master_parsed_df["Col"])
    master_df.insert(7, "Width", master_parsed_df["Width"])
    master_df.insert(8, "Position", master_parsed_df["Position"])

    operation_summary_df = (
        master_df.groupby(["Operation Name", "Style", "Color"], sort=False, dropna=False)["PLM No"]
        .apply(lambda series: series.replace("", pd.NA).count())
        .reset_index(name="PLM No Count")
    )

    color_order = merged_data["Color"].dropna().astype(str).drop_duplicates().tolist()
    combine_operation_series = combine_df["Operation Name"].fillna("").astype(str).str.strip()
    combine_plm_series = combine_df["PLM No"].fillna("").astype(str).str.strip()
    combine_non_blank_plm_mask = combine_plm_series != ""

    operation_row_indices: dict[str, list[int]] = {}
    combine_non_blank_df = pd.DataFrame(
        {
            "Row Index": combine_df.index,
            "Operation Name": combine_operation_series,
            "PLM No": combine_plm_series,
        }
    )
    combine_non_blank_df = combine_non_blank_df[combine_non_blank_plm_mask]
    combine_non_blank_df = combine_non_blank_df.sort_values(["Operation Name", "PLM No"], kind="stable")

    for _, row in combine_non_blank_df.iterrows():
        operation_name = str(row["Operation Name"])
        row_index = int(row["Row Index"])
        operation_row_indices.setdefault(operation_name, []).append(row_index)

    master_operation_series = master_df["Operation Name"].fillna("").astype(str).str.strip()
    master_color_series = master_df["Color"].fillna("").astype(str).str.strip()
    master_col_series = master_df["Col"].fillna("").astype(str).str.strip()

    for color in color_order:
        if color not in combine_df.columns:
            combine_df[color] = ""

        for operation_name, row_indices in operation_row_indices.items():
            master_match_mask = (
                (master_operation_series == operation_name)
                & (master_color_series == color)
                & (master_col_series != "")
            )
            col_values = sorted(master_col_series[master_match_mask].tolist())

            for row_index, col_value in zip(row_indices, col_values):
                combine_df.at[row_index, color] = col_value

    combine_df = combine_df.drop(columns=["Color"], errors="ignore")

    combine_lookup = {
        (normalize_key(plm_no), normalize_position_key(position_value))
        for plm_no, position_value in zip(combine_df["PLM No"], combine_df["Position"])
        if normalize_key(plm_no)
    }
    combine_plm_lookup = {
        normalize_key(plm_no)
        for plm_no in combine_df["PLM No"]
        if normalize_key(plm_no)
    }

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        master_df.to_excel(writer, index=False, sheet_name="Master")
        operation_summary_df.to_excel(writer, index=False, sheet_name="Operation Summary")
        item_calc_df.to_excel(writer, index=False, sheet_name="Item Calculation")
        combine_df.to_excel(writer, index=False, sheet_name="Combine")

        if plm_df is not None:
            plm_output_df = plm_df.copy()
            if not plm_output_df.empty and plm_output_df.shape[1] >= 2:
                plm_output_df.iat[0, 1] = "PLM No"

            if not plm_output_df.empty and plm_no_index is not None and plm_position_index is not None:
                plm_check_values = ["PLM Check"]
                for _, row in plm_output_df.iloc[1:].iterrows():
                    plm_raw = "" if pd.isna(row.iloc[plm_no_index]) else str(row.iloc[plm_no_index]).strip()
                    plm_prefix = plm_raw[:1]
                    plm_key = normalize_key(plm_raw)
                    position_key = normalize_position_key(row.iloc[plm_position_index])

                    if plm_prefix in {"R", "S"}:
                        is_match = bool(plm_key and (plm_key, position_key) in combine_lookup)
                    elif plm_prefix == "P":
                        is_match = bool(plm_key and plm_key in combine_plm_lookup)
                    else:
                        is_match = False

                    plm_check_values.append("Yes" if is_match else "No")

                plm_output_df.insert(0, "PLM Check", plm_check_values)
            else:
                logs.append("PLM Check warning: Could not create PLM Check column due to missing PLM columns.")

            plm_output_df.to_excel(writer, index=False, header=False, sheet_name="PLM")
            format_output_sheet(writer.book["PLM"], highlight_plm_check_header=True)

        format_output_sheet(writer.book["Master"], highlight_operation_check_yes=True)
        format_output_sheet(
            writer.book["Item Calculation"],
            highlight_row_column_name="Color",
            highlight_row_match_value=selected_color,
        )
        format_output_sheet(writer.book["Operation Summary"])
        format_output_sheet(writer.book["Combine"], highlight_numeric_operation=True, freeze_top_row=True)

    output.seek(0)
    previews = {
        "Master": master_df,
        "Operation Summary": operation_summary_df,
        "Item Calculation": item_calc_df,
        "Combine": combine_df,
    }
    return output.getvalue(), previews, logs


def main() -> None:
    st.set_page_config(page_title="Consumption Saving", layout="wide")
    apply_app_styles()
    st.markdown(
        """
        <div class="hero-card">
            <div class="hero-title">Consumption Calculation - Triumph Vietnam</div>
            <div class="hero-subtitle">Upload source Excel files and optional PLM file, then generate a merged workbook in one click.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="section-title">Input Configuration</div>', unsafe_allow_html=True)
    left_col, right_col = st.columns([2, 1], gap="large")
    with left_col:
        style_name = st.text_input("Style name *", placeholder="Enter style name")
        source_files = st.file_uploader(
            "Upload source Excel files",
            type=["xlsx", "xls", "xlsm", "xlsb"],
            accept_multiple_files=True,
        )
    with right_col:
        plm_file = st.file_uploader(
            "Upload PLM file (optional) - Maximum 1 file",
            type=["xlsx", "xls", "xlsm", "xlsb"],
            accept_multiple_files=False,
        )

    source_count = len(source_files) if source_files else 0
    style_ready = bool(style_name.strip())
    st.markdown(
        (
            '<div class="status-wrap">'
            + (f'<span class="status-pill status-ok">{source_count} source file(s) selected</span>' if source_count else '<span class="status-pill status-warn">No source files selected</span>')
            + ('<span class="status-pill status-ok">Style name ready</span>' if style_ready else '<span class="status-pill status-warn">Style name is required</span>')
            + (f'<span class="status-pill status-ok">PLM: {plm_file.name}</span>' if plm_file else '<span class="status-pill status-warn">PLM file not uploaded</span>')
            + '</div>'
        ),
        unsafe_allow_html=True,
    )

    run_clicked = st.button(
        "Generate Output",
        type="primary",
        disabled=not style_name.strip() or not source_files,
        use_container_width=True,
    )

    if run_clicked:
        if not style_name.strip():
            st.error("Please enter style name.")
            return
        if not source_files:
            st.error("Please upload at least one source Excel file.")
            return

        with st.spinner("Processing files..."):
            try:
                merged_df, summary, read_logs = build_master_from_uploads(source_files, style_name.strip())
                output_bytes, previews, build_logs = generate_output_workbook(
                    merged_df,
                    style_name.strip(),
                    summary,
                    plm_file,
                )
            except Exception as e:
                st.error(f"Processing failed: {e}")
                return

        st.success("Completed successfully.")

        metric_col1, metric_col2, metric_col3 = st.columns(3)
        with metric_col1:
            st.metric("Source Files", len(source_files))
        with metric_col2:
            st.metric("Master Rows", len(previews["Master"]))
        with metric_col3:
            st.metric("PLM Included", "Yes" if plm_file else "No")

        log_lines = read_logs + build_logs
        if log_lines:
            st.subheader("Logs")
            st.text("\n".join(log_lines))

        st.subheader("Preview - Master (Top 100 rows)")
        st.dataframe(previews["Master"].head(100), use_container_width=True)

        st.download_button(
            label="Download merged workbook",
            data=output_bytes,
            file_name=f"{style_name.strip()}_merged.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
